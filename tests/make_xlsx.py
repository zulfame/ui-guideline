"""Create test xlsx files for import UI testing."""
import openpyxl
import os
import time

OUT = "/tmp/import_tests"
os.makedirs(OUT, exist_ok=True)
TS = int(time.time())


def save(wb, name):
    p = f"{OUT}/{name}"
    wb.save(p)
    print(p)
    return p


# ---------- Offices happy path ----------
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["code", "name", "address", "longitude", "latitude", "radius", "note"])
ws.append([f"ZZ{TS}A", f"ZZ Office A {TS}", "Addr A", 106.8, -6.2, 100, "note A"])
ws.append([f"ZZ{TS}B", f"ZZ Office B {TS}", "Addr B", 107.6, -6.9, 150, "note B"])
save(wb, "offices_ok.xlsx")

# ---------- Offices error path (row 3 missing name, longitude=999) ----------
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["code", "name", "address", "longitude", "latitude", "radius", "note"])
ws.append([f"ZZ{TS}E1", f"ZZ Office E1 {TS}", "Addr E1", 106.8, -6.2, 100, ""])
ws.append([f"ZZ{TS}E2", "", "Addr E2", 999, -6.9, 150, ""])
save(wb, "offices_bad.xlsx")

# ---------- Roles happy path (row2 parent references row1) ----------
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["name", "parent", "level", "order"])
ws.append([f"ZZ Boss {TS}", "", "", 0])
ws.append([f"ZZ Sub {TS}", f"ZZ Boss {TS}", "", 0])
save(wb, "roles_ok.xlsx")

# ---------- Users happy path ----------
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["name", "email", "role", "office", "username"])
ws.append([f"ZZ User {TS}", f"zzuser{TS}@example.com", "Chief Executive Officer", "Head Office", f"zzu{TS}"])
save(wb, "users_ok.xlsx")

# ---------- Users error path (bad role name) ----------
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["name", "email", "role", "office", "username"])
ws.append([f"ZZ UserBad {TS}", f"zzuserbad{TS}@example.com", "NOPE-ROLE-XYZ", "Head Office", f"zzub{TS}"])
save(wb, "users_bad.xlsx")

print(f"TS={TS}")
