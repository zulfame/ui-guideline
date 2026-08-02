"""Excel Import domain — Offices / Roles / Users template download + bulk upsert.

Extracted from server.py (behavior unchanged). Routes register on the shared
`api_router` at import time. All-or-nothing validation with upsert on the natural
key (code / name / email); reference columns use human-readable names -> UUIDs.
"""
import io
import re
import uuid
from datetime import datetime, timedelta, timezone
from typing import List, Optional

from fastapi import File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from pymongo import InsertOne, UpdateOne

from server import (
    api_router,
    db,
    log_audit,
    _hash_password,
    _next_user_id,
    DEFAULT_USER_PASSWORD,
    PASSWORD_EXPIRY_DAYS,
    UNIQUE_USER_FIELDS,
    EMAIL_RE,
)


# ---------------------------------------------------------------------------
# Excel Import (Offices / Roles / Users) — template download + bulk upsert.
# Policy: all-or-nothing (validate every row first; abort on any error and
# report the offending rows) with upsert on the natural key (code / name /
# email). Reference columns use human-readable names, resolved to UUIDs.
# ---------------------------------------------------------------------------
def _s(v) -> str:
    """Coerce an Excel cell to a trimmed string (ints stay clean, no '.0')."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _to_float(v, field, errs):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        errs.append(f"{field} must be a number")
        return None


def _to_int_or_none(v):
    """Return int, 0 for blank, or None when the value is not a whole number."""
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _read_upload(content: bytes) -> list:
    """Parse the first worksheet into a list of dicts keyed by lower-cased headers."""
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail={"message": "Could not read the file. Please upload a valid .xlsx Excel file.", "errors": []},
        )
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [(_s(h).lower() if h is not None else "") for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None or all(c is None or (isinstance(c, str) and c.strip() == "") for c in r):
            continue
        d = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            d[h] = r[idx] if idx < len(r) else None
        out.append(d)
    return out


def _xlsx_response(headers: list, example: list, filename: str) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(headers)
    if example:
        ws.append(example)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _abort_import(errors: list):
    raise HTTPException(
        status_code=400,
        detail={
            "message": f"Import canceled: {len(errors)} row(s) have problems. Fix them and try again.",
            "errors": errors,
        },
    )


# UI-friendly cap: never stream thousands of preview rows to the browser.
PREVIEW_ROW_CAP = 300


def _preview_response(errors: list, items: list) -> dict:
    to_create = sum(1 for i in items if i["action"] == "create")
    to_update = sum(1 for i in items if i["action"] == "update")
    preview = [
        {"row": i["row"], "action": i["action"], "label": i["label"]}
        for i in items[:PREVIEW_ROW_CAP]
    ]
    return {
        "total": len(items),
        "to_create": to_create,
        "to_update": to_update,
        "errors": errors,
        "rows": preview,
        "truncated": len(items) > PREVIEW_ROW_CAP,
    }


class ImportErrorsExport(BaseModel):
    errors: List[dict] = []
    filename: Optional[str] = None


@api_router.post("/import/errors/export", tags=["Import"], summary="Export import errors to Excel")
async def export_import_errors(payload: ImportErrorsExport):
    """Return the failing rows as an .xlsx (columns: Row, Errors) for easy fixing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"
    ws.append(["Row", "Errors"])
    for e in payload.errors:
        ws.append([e.get("row"), "; ".join(e.get("errors", []))])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = payload.filename or "import_errors.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---- Offices ----
@api_router.get("/offices/import/template", tags=["Offices"], summary="Download offices import template")
async def offices_import_template():
    headers = ["code", "name", "address", "telephone", "longitude", "latitude", "radius", "note"]
    example = ["KP001", "Kantor Pusat", "Jl. Sudirman No. 1", "021-5550000", 106.8272, -6.2088, 100, "Catatan opsional"]
    return _xlsx_response(headers, example, "offices_import_template.xlsx")


@api_router.post("/offices/import", tags=["Offices"], summary="Import offices from Excel")
async def import_offices(file: UploadFile = File(...)):
    rows = _read_upload(await file.read())
    if not rows:
        raise HTTPException(status_code=400, detail={"message": "No data rows found in the file.", "errors": []})
    errors, items = await _prepare_offices(rows)
    if errors:
        _abort_import(errors)
    created, updated = await _apply_offices(items)
    await log_audit(
        "import", "office",
        summary=f"Imported offices: {created} created, {updated} updated",
        method="POST", path="/api/offices/import", status_code=200,
        request={"filename": file.filename, "rows": len(items)},
        response={"created": created, "updated": updated, "total": len(items)},
        metadata={"created": created, "updated": updated, "total": len(items)},
    )
    return {"success": True, "created": created, "updated": updated, "total": len(items)}


@api_router.post("/offices/import/preview", tags=["Offices"], summary="Preview offices import (dry-run)")
async def preview_offices(file: UploadFile = File(...)):
    rows = _read_upload(await file.read())
    if not rows:
        raise HTTPException(status_code=400, detail={"message": "No data rows found in the file.", "errors": []})
    errors, items = await _prepare_offices(rows)
    return _preview_response(errors, items)


async def _prepare_offices(rows: list):
    errors, parsed = [], []
    seen_code, seen_name = {}, {}
    for i, row in enumerate(rows):
        rn = i + 2
        errs = []
        code = _s(row.get("code"))
        name = _s(row.get("name"))
        if not code:
            errs.append("code is required")
        if not name:
            errs.append("name is required")
        if code:
            if code.lower() in seen_code:
                errs.append(f"duplicate code in file (row {seen_code[code.lower()]})")
            else:
                seen_code[code.lower()] = rn
        if name:
            if name.lower() in seen_name:
                errs.append(f"duplicate name in file (row {seen_name[name.lower()]})")
            else:
                seen_name[name.lower()] = rn
        lon = _to_float(row.get("longitude"), "longitude", errs)
        lat = _to_float(row.get("latitude"), "latitude", errs)
        rad = _to_float(row.get("radius"), "radius", errs)
        if lon is not None and not (-180 <= lon <= 180):
            errs.append("longitude must be between -180 and 180")
        if lat is not None and not (-90 <= lat <= 90):
            errs.append("latitude must be between -90 and 90")
        parsed.append({
            "row": rn, "code": code, "name": name,
            "address": _s(row.get("address")) or None,
            "telephone": _s(row.get("telephone")) or None,
            "longitude": lon, "latitude": lat,
            "radius": rad if rad is not None else 100,
            "note": _s(row.get("note")) or None,
        })
        if errs:
            errors.append({"row": rn, "errors": errs})
    existing = await db.offices.find({}, {"_id": 0, "id": 1, "code": 1, "name": 1}).to_list(100000)
    by_code = {o["code"].lower(): o for o in existing}
    name_owner = {o["name"].lower(): o["code"].lower() for o in existing}
    for p in parsed:
        if not p["name"] or not p["code"]:
            continue
        owner = name_owner.get(p["name"].lower())
        if owner is not None and owner != p["code"].lower():
            errors.append({"row": p["row"], "errors": [f"name already used by another office (code {owner.upper()})"]})
    items = []
    for p in parsed:
        ex = by_code.get(p["code"].lower())
        items.append({
            "row": p["row"],
            "action": "update" if ex else "create",
            "label": f"{p['code']} — {p['name']}",
            "id": ex["id"] if ex else None,
            "fields": {
                "code": p["code"], "name": p["name"], "address": p["address"],
                "telephone": p["telephone"], "longitude": p["longitude"],
                "latitude": p["latitude"], "radius": p["radius"], "note": p["note"],
            },
        })
    return errors, items


async def _apply_offices(items: list):
    now = datetime.now(timezone.utc).isoformat()
    ops, created, updated = [], 0, 0
    for it in items:
        fields = {**it["fields"], "updated_at": now}
        if it["id"]:
            ops.append(UpdateOne({"id": it["id"]}, {"$set": fields}))
            updated += 1
        else:
            ops.append(InsertOne({"id": str(uuid.uuid4()), **fields, "created_at": now}))
            created += 1
    if ops:
        await db.offices.bulk_write(ops, ordered=False)
    return created, updated


# ---- Roles ----
@api_router.get("/roles/import/template", tags=["Roles"], summary="Download roles import template")
async def roles_import_template():
    headers = ["name", "parent", "dotted_parent", "level", "order"]
    example = ["Teller", "Kepala Cabang", "", "Staff", 0]
    return _xlsx_response(headers, example, "roles_import_template.xlsx")


@api_router.post("/roles/import", tags=["Roles"], summary="Import roles from Excel")
async def import_roles(file: UploadFile = File(...)):
    rows = _read_upload(await file.read())
    if not rows:
        raise HTTPException(status_code=400, detail={"message": "No data rows found in the file.", "errors": []})
    errors, items = await _prepare_roles(rows)
    if errors:
        _abort_import(errors)
    created, updated = await _apply_roles(items)
    await log_audit(
        "import", "role",
        summary=f"Imported roles: {created} created, {updated} updated",
        method="POST", path="/api/roles/import", status_code=200,
        request={"filename": file.filename, "rows": len(items)},
        response={"created": created, "updated": updated, "total": len(items)},
        metadata={"created": created, "updated": updated, "total": len(items)},
    )
    return {"success": True, "created": created, "updated": updated, "total": len(items)}


@api_router.post("/roles/import/preview", tags=["Roles"], summary="Preview roles import (dry-run)")
async def preview_roles(file: UploadFile = File(...)):
    rows = _read_upload(await file.read())
    if not rows:
        raise HTTPException(status_code=400, detail={"message": "No data rows found in the file.", "errors": []})
    errors, items = await _prepare_roles(rows)
    return _preview_response(errors, items)


async def _prepare_roles(rows: list):
    existing = await db.roles.find({}, {"_id": 0, "id": 1, "name": 1, "parent_id": 1}).to_list(100000)
    levels = await db.levels.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100000)
    role_id_by_name = {r["name"].lower(): r["id"] for r in existing}
    level_id_by_name = {l["name"].lower(): l["id"] for l in levels}
    errors, parsed, seen_name, file_names = [], [], {}, set()
    for i, row in enumerate(rows):
        rn = i + 2
        errs = []
        name = _s(row.get("name"))
        if not name:
            errs.append("name is required")
        if name:
            if name.lower() in seen_name:
                errs.append(f"duplicate name in file (row {seen_name[name.lower()]})")
            else:
                seen_name[name.lower()] = rn
                file_names.add(name.lower())
        order = _to_int_or_none(row.get("order"))
        if order is None:
            errs.append("order must be a whole number")
            order = 0
        parsed.append({
            "row": rn, "name": name,
            "parent": _s(row.get("parent")),
            "dotted": _s(row.get("dotted_parent")),
            "level": _s(row.get("level")),
            "order": order,
        })
        if errs:
            errors.append({"row": rn, "errors": errs})
    all_names = set(role_id_by_name) | file_names
    for p in parsed:
        errs = []
        if p["parent"]:
            if p["parent"].lower() == p["name"].lower():
                errs.append("parent cannot be the role itself")
            elif p["parent"].lower() not in all_names:
                errs.append(f"parent role '{p['parent']}' not found")
        if p["dotted"]:
            if p["dotted"].lower() == p["name"].lower():
                errs.append("dotted_parent cannot be the role itself")
            elif p["dotted"].lower() not in all_names:
                errs.append(f"dotted_parent role '{p['dotted']}' not found")
        if p["level"] and p["level"].lower() not in level_id_by_name:
            errs.append(f"level '{p['level']}' not found")
        if errs:
            errors.append({"row": p["row"], "errors": errs})
    name_to_id = dict(role_id_by_name)
    for p in parsed:
        if p["name"]:
            name_to_id.setdefault(p["name"].lower(), str(uuid.uuid4()))
    final_parent = {r["id"]: r.get("parent_id") for r in existing}
    for p in parsed:
        if not p["name"]:
            continue
        rid = name_to_id[p["name"].lower()]
        final_parent[rid] = name_to_id.get(p["parent"].lower()) if p["parent"] else None

    def _has_cycle(start):
        seen, cur = set(), start
        while cur is not None:
            if cur in seen:
                return True
            seen.add(cur)
            cur = final_parent.get(cur)
        return False

    if not errors:  # cycle check only meaningful once names resolve cleanly
        cyc = [p["row"] for p in parsed if p["name"] and _has_cycle(name_to_id[p["name"].lower()])]
        for r in cyc:
            errors.append({"row": r, "errors": ["parent relationships create a cycle"]})
    existing_ids = {r["id"] for r in existing}
    items = []
    for p in parsed:
        if not p["name"]:
            continue
        rid = name_to_id[p["name"].lower()]
        items.append({
            "row": p["row"],
            "action": "update" if rid in existing_ids else "create",
            "label": p["name"],
            "id": rid,
            "is_new": rid not in existing_ids,
            "fields": {
                "name": p["name"],
                "parent_id": name_to_id.get(p["parent"].lower()) if p["parent"] else None,
                "dotted_parent_id": name_to_id.get(p["dotted"].lower()) if p["dotted"] else None,
                "level_id": level_id_by_name.get(p["level"].lower()) if p["level"] else None,
                "order": p["order"],
            },
        })
    return errors, items


async def _apply_roles(items: list):
    now = datetime.now(timezone.utc).isoformat()
    ops, created, updated = [], 0, 0
    for it in items:
        fields = {**it["fields"], "updated_at": now}
        if it["is_new"]:
            ops.append(InsertOne({"id": it["id"], **fields, "created_at": now}))
            created += 1
        else:
            ops.append(UpdateOne({"id": it["id"]}, {"$set": fields}))
            updated += 1
    if ops:
        await db.roles.bulk_write(ops, ordered=False)
    return created, updated


# ---- Users ----
_USER_IMPORT_OPTIONAL = [
    "username", "phone", "alias", "mso_code", "collector_code",
    "device_identifier", "device_name", "device_os", "fcm_token",
]


def _parse_active(val) -> bool:
    """Parse the import 'is_active'/'status' cell → bool (default active)."""
    s = _s(val).lower()
    if s == "":
        return True
    return s not in ("0", "false", "no", "n", "inactive", "nonaktif", "tidak aktif", "tidak", "off")


@api_router.get("/users/import/template", tags=["Users"], summary="Download users import template")
async def users_import_template():
    headers = ["user_id", "name", "email", "role", "office", "is_active", *_USER_IMPORT_OPTIONAL]
    example = [
        "", "Budi Santoso", "budi@example.com", "Teller", "Kantor Pusat", "active",
        "budi", "08123456789", "BDS", "MSO001", "COL001", "", "", "", "",
    ]
    return _xlsx_response(headers, example, "users_import_template.xlsx")


@api_router.post("/users/import", tags=["Users"], summary="Import users from Excel")
async def import_users(file: UploadFile = File(...)):
    rows = _read_upload(await file.read())
    if not rows:
        raise HTTPException(status_code=400, detail={"message": "No data rows found in the file.", "errors": []})
    errors, items = await _prepare_users(rows)
    if errors:
        _abort_import(errors)
    created, updated = await _apply_users(items)
    await log_audit(
        "import", "user",
        summary=f"Imported users: {created} created, {updated} updated",
        method="POST", path="/api/users/import", status_code=200,
        request={"filename": file.filename, "rows": len(items)},
        response={"created": created, "updated": updated, "total": len(items)},
        metadata={"created": created, "updated": updated, "total": len(items)},
    )
    return {"success": True, "created": created, "updated": updated, "total": len(items)}


@api_router.post("/users/import/preview", tags=["Users"], summary="Preview users import (dry-run)")
async def preview_users(file: UploadFile = File(...)):
    rows = _read_upload(await file.read())
    if not rows:
        raise HTTPException(status_code=400, detail={"message": "No data rows found in the file.", "errors": []})
    errors, items = await _prepare_users(rows)
    return _preview_response(errors, items)


async def _prepare_users(rows: list):
    roles = await db.roles.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100000)
    offices = await db.offices.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100000)
    role_id_by_name = {r["name"].lower(): r["id"] for r in roles}
    office_id_by_name = {o["name"].lower(): o["id"] for o in offices}
    existing_users = await db.users.find({"deleted_at": None}, {"_id": 0}).to_list(1000000)
    user_by_email = {u["email"].lower(): u for u in existing_users if u.get("email")}
    errors, parsed = [], []
    seen = {f: {} for f in UNIQUE_USER_FIELDS}
    for i, row in enumerate(rows):
        rn = i + 2
        errs = []
        name = _s(row.get("name"))
        email = _s(row.get("email"))
        role_name = _s(row.get("role"))
        office_name = _s(row.get("office"))
        if not name:
            errs.append("name is required")
        if not email:
            errs.append("email is required")
        elif not re.match(EMAIL_RE, email):
            errs.append("email format is invalid")
        role_id = office_id = None
        if not role_name:
            errs.append("role is required")
        elif role_name.lower() not in role_id_by_name:
            errs.append(f"role '{role_name}' not found")
        else:
            role_id = role_id_by_name[role_name.lower()]
        if not office_name:
            errs.append("office is required")
        elif office_name.lower() not in office_id_by_name:
            errs.append(f"office '{office_name}' not found")
        else:
            office_id = office_id_by_name[office_name.lower()]
        rec = {"row": rn, "name": name, "email": email, "role_id": role_id, "office_id": office_id}
        uid_raw = _s(row.get("user_id"))
        user_id_val = None
        if uid_raw:
            if re.fullmatch(r"\d+", uid_raw) and int(uid_raw) >= 1:
                user_id_val = int(uid_raw)
            else:
                errs.append("user_id must be a positive whole number")
        rec["user_id"] = user_id_val
        rec["is_active"] = _parse_active(row.get("is_active", row.get("status")))
        for f in _USER_IMPORT_OPTIONAL:
            rec[f] = _s(row.get(f)) or None
        for f in UNIQUE_USER_FIELDS:
            val = email if f == "email" else rec.get(f)
            if val:
                key = val.lower() if f == "email" else val
                if key in seen[f]:
                    errs.append(f"duplicate {f} in file (row {seen[f][key]})")
                else:
                    seen[f][key] = rn
        parsed.append(rec)
        if errs:
            errors.append({"row": rn, "errors": errs})
    for p in parsed:
        target = user_by_email.get(p["email"].lower()) if p["email"] else None
        target_id = target["id"] if target else None
        errs = []
        for f in UNIQUE_USER_FIELDS:
            val = p["email"] if f == "email" else p.get(f)
            if not val:
                continue
            q = {f: val, "deleted_at": None}
            if target_id:
                q = {"$and": [{"id": {"$ne": target_id}}, q]}
            if await db.users.find_one(q, {"_id": 0, "id": 1}):
                errs.append(f"{f} already exists for another user")
        if errs:
            errors.append({"row": p["row"], "errors": errs})
    items = []
    for p in parsed:
        target = user_by_email.get(p["email"].lower()) if p["email"] else None
        items.append({
            "row": p["row"],
            "action": "update" if target else "create",
            "label": f"{p['name']} <{p['email']}>",
            "target_id": target["id"] if target else None,
            "fields": {
                "name": p["name"], "email": p["email"],
                "role_id": p["role_id"], "office_id": p["office_id"],
                "is_active": p["is_active"], "user_id": p["user_id"],
                **{f: p[f] for f in _USER_IMPORT_OPTIONAL},
            },
        })
    return errors, items


async def _apply_users(items: list):
    now = datetime.now(timezone.utc)
    now_iso = now.isoformat()
    created, updated = 0, 0
    for it in items:
        base = {**it["fields"], "updated_at": now_iso}
        provided_uid = base.pop("user_id", None)
        if it["target_id"]:
            if provided_uid is not None:
                base["user_id"] = provided_uid
            await db.users.update_one({"id": it["target_id"]}, {"$set": base})
            updated += 1
        else:
            uid = provided_uid if provided_uid is not None else await _next_user_id()
            pw = _hash_password(DEFAULT_USER_PASSWORD)
            await db.users.insert_one({
                "id": str(uuid.uuid4()), "user_id": uid, **base,
                "password": pw, "password_history": [pw],
                "password_changed_at": now_iso,
                "password_expires_at": (now + timedelta(days=PASSWORD_EXPIRY_DAYS)).isoformat(),
                "must_change_password": True, "deleted_at": None, "created_at": now_iso,
            })
            created += 1
    return created, updated
